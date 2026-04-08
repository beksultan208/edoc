"""
ГосДок — Генераторы отчётов (apps/reports/generators.py)
Раздел 2.9 ТЗ: PDF (WeasyPrint) и XLSX (openpyxl) экспорт.

Формат отчёта (раздел 3.11 ТЗ):
  - Заголовок организации и кабинета
  - Период (год/месяц)
  - Таблица показателей: docs_total, docs_completed, docs_signed,
    tasks_completed, avg_completion_days
"""

import io
import logging
from datetime import date

logger = logging.getLogger(__name__)


# ============================================================
# Сбор статистики за период
# ============================================================

def generate_report_data_by_workspace(workspace_id: str, year: int, month: int) -> dict:
    """
    Собирает статистику кабинета за указанный месяц.

    Подсчёт ведётся по:
    - Document: created_at в диапазоне [start_date, end_date)
    - Task: completed_at в том же диапазоне, status='done'
    - avg_completion_days: среднее (updated_at − created_at) для signed-документов

    Returns:
        dict с полями модели MonthlyReport + ключ 'report_data' (JSONB)
    """
    from apps.documents.models import Document
    from apps.tasks.models import Task

    # Границы периода
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    docs = Document.objects.filter(
        workspace_id=workspace_id,
        created_at__date__gte=start_date,
        created_at__date__lt=end_date,
    )

    docs_total = docs.count()
    docs_completed = docs.filter(status__in=["signed", "archived"]).count()
    docs_signed = docs.filter(status="signed").count()

    tasks_completed = Task.objects.filter(
        workspace_id=workspace_id,
        status="done",
        completed_at__date__gte=start_date,
        completed_at__date__lt=end_date,
    ).count()

    # Среднее время согласования (дней): среднее по подписанным документам
    avg_days = None
    signed_docs = docs.filter(status="signed", updated_at__isnull=False)
    if signed_docs.exists():
        total_days = sum(
            (d.updated_at.date() - d.created_at.date()).days
            for d in signed_docs
        )
        avg_days = round(total_days / signed_docs.count(), 2)

    return {
        "docs_total": docs_total,
        "docs_completed": docs_completed,
        "docs_signed": docs_signed,
        "tasks_completed": tasks_completed,
        "avg_completion_days": avg_days,
        "report_data": {
            "period": f"{year}-{month:02d}",
            "workspace_id": str(workspace_id),
        },
    }


# ============================================================
# Вспомогательные функции
# ============================================================

def _get_org_name(report) -> str:
    """
    Возвращает название организации для отчёта.
    Приоритет: report.organization → report.workspace.organization → '—'
    """
    if report.organization:
        return report.organization.name
    if report.workspace and report.workspace.organization:
        return report.workspace.organization.name
    return "—"


def _get_workspace_title(report) -> str:
    """Возвращает название кабинета или '—'."""
    return report.workspace.title if report.workspace else "—"


# ============================================================
# Экспорт в XLSX
# ============================================================

def export_report_to_xlsx(report) -> bytes:
    """
    Экспортирует отчёт в XLSX (опenpyxl).
    Раздел 2.9 ТЗ: экспорт ежемесячного отчёта.

    Структура файла:
    - Лист 1 «Показатели»: сводная таблица по разделу 3.11 ТЗ
    - Лист 2 «Инфо»: метаданные отчёта

    Returns:
        bytes — содержимое .xlsx-файла
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ---- Лист 1: Показатели ----
        ws = wb.active
        ws.title = f"{report.period_month:02d}.{report.period_year}"

        period_str = f"{report.period_month:02d}/{report.period_year}"
        org_name = _get_org_name(report)
        workspace_title = _get_workspace_title(report)

        # Заголовок
        ws["A1"] = "ГосДок — Ежемесячный отчёт"
        ws["A1"].font = Font(bold=True, size=16, color="2C3E50")
        ws.merge_cells("A1:C1")

        ws["A2"] = f"Организация: {org_name}"
        ws["A2"].font = Font(bold=True, size=12)
        ws.merge_cells("A2:C2")

        ws["A3"] = f"Кабинет: {workspace_title}"
        ws["A3"].font = Font(size=11)
        ws.merge_cells("A3:C3")

        ws["A4"] = f"Период: {period_str}"
        ws["A4"].font = Font(size=11)
        ws.merge_cells("A4:C4")

        # Разделитель
        ws["A5"] = ""

        # Шапка таблицы
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        headers = [("Показатель", 40), ("Значение", 20), ("Примечание", 30)]
        for col_idx, (header, width) in enumerate(headers, start=1):
            cell = ws.cell(row=6, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Данные
        rows_data = [
            ("Всего документов за период", report.docs_total, ""),
            ("Завершённых документов (подписаны + архив)", report.docs_completed, ""),
            ("Подписанных документов", report.docs_signed, ""),
            ("Завершённых задач workflow", report.tasks_completed, ""),
            (
                "Среднее время согласования (дней)",
                float(report.avg_completion_days) if report.avg_completion_days is not None else "—",
                "calc. от создания до подписи",
            ),
        ]

        alt_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
        for row_idx, (label, value, note) in enumerate(rows_data, start=7):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=3, value=note)
            if row_idx % 2 == 0:
                for col in range(1, 4):
                    ws.cell(row=row_idx, column=col).fill = alt_fill

        # Дата генерации
        ws.cell(row=13, column=1, value=f"Сгенерировано: {report.generated_at.strftime('%d.%m.%Y %H:%M')}")
        ws.cell(row=13, column=1).font = Font(italic=True, color="7F8C8D", size=9)

        # ---- Лист 2: Информация ----
        ws2 = wb.create_sheet(title="Инфо")
        ws2["A1"] = "Поле"
        ws2["B1"] = "Значение"
        ws2["A1"].font = Font(bold=True)
        ws2["B1"].font = Font(bold=True)

        meta_rows = [
            ("ID отчёта", str(report.id)),
            ("Организация", org_name),
            ("Кабинет", workspace_title),
            ("Год", report.period_year),
            ("Месяц", report.period_month),
            ("Дата генерации", report.generated_at.strftime("%d.%m.%Y %H:%M")),
        ]
        for r, (key, val) in enumerate(meta_rows, start=2):
            ws2.cell(row=r, column=1, value=key)
            ws2.cell(row=r, column=2, value=val)
        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 40

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    except ImportError:
        logger.error("openpyxl не установлен — экспорт XLSX недоступен")
        return b""
    except Exception as exc:
        logger.error("Ошибка при генерации XLSX: %s", exc)
        return b""


# ============================================================
# Экспорт в PDF
# ============================================================

def export_report_to_pdf(report) -> bytes:
    """
    Экспортирует отчёт в PDF с помощью WeasyPrint.
    Раздел 2.9 ТЗ: «заголовок организации, период, таблица показателей».

    HTML-шаблон включает:
    - Логотип/заголовок «ГосДок»
    - Название организации (раздел 3.11 ТЗ)
    - Название кабинета
    - Период
    - Таблицу всех показателей из раздела 3.11 ТЗ

    Returns:
        bytes — содержимое .pdf-файла
    """
    try:
        import weasyprint

        org_name = _get_org_name(report)
        workspace_title = _get_workspace_title(report)
        period_str = f"{report.period_month:02d}/{report.period_year}"
        generated_str = report.generated_at.strftime("%d.%m.%Y %H:%M")
        avg_days = (
            f"{float(report.avg_completion_days):.2f} дн."
            if report.avg_completion_days is not None
            else "—"
        )

        html_content = f"""
        <!DOCTYPE html>
        <html lang="ru">
        <head>
            <meta charset="utf-8">
            <style>
                @page {{
                    size: A4;
                    margin: 20mm 15mm;
                }}
                body {{
                    font-family: "Arial", "DejaVu Sans", sans-serif;
                    font-size: 11pt;
                    color: #2c3e50;
                    padding: 0;
                    margin: 0;
                }}
                .header {{
                    border-bottom: 3px solid #3498db;
                    padding-bottom: 12px;
                    margin-bottom: 20px;
                }}
                .header h1 {{
                    font-size: 22pt;
                    color: #2c3e50;
                    margin: 0 0 4px 0;
                }}
                .header .subtitle {{
                    font-size: 10pt;
                    color: #7f8c8d;
                    margin: 0;
                }}
                .meta-block {{
                    background: #f8f9fa;
                    border-left: 4px solid #3498db;
                    padding: 10px 16px;
                    margin-bottom: 20px;
                    border-radius: 0 4px 4px 0;
                }}
                .meta-block p {{
                    margin: 4px 0;
                    font-size: 11pt;
                }}
                .meta-block .label {{
                    font-weight: bold;
                    color: #555;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 10px;
                }}
                thead tr {{
                    background-color: #3498db;
                    color: white;
                }}
                thead th {{
                    padding: 10px 12px;
                    text-align: left;
                    font-size: 11pt;
                }}
                tbody tr:nth-child(even) {{
                    background-color: #ecf0f1;
                }}
                tbody td {{
                    padding: 9px 12px;
                    border-bottom: 1px solid #dce1e7;
                    font-size: 11pt;
                }}
                tbody td:last-child {{
                    text-align: center;
                    font-weight: bold;
                    color: #2980b9;
                }}
                .footer {{
                    margin-top: 30px;
                    border-top: 1px solid #dce1e7;
                    padding-top: 8px;
                    font-size: 9pt;
                    color: #95a5a6;
                    text-align: right;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>ГосДок — Ежемесячный отчёт</h1>
                <p class="subtitle">Облачная платформа документооборота</p>
            </div>

            <div class="meta-block">
                <p><span class="label">Организация:</span> {org_name}</p>
                <p><span class="label">Кабинет:</span> {workspace_title}</p>
                <p><span class="label">Период:</span> {period_str}</p>
            </div>

            <table>
                <thead>
                    <tr>
                        <th>Показатель</th>
                        <th>Значение</th>
                    </tr>
                </thead>
                <tbody>
                    <tr>
                        <td>Всего документов за период</td>
                        <td>{report.docs_total}</td>
                    </tr>
                    <tr>
                        <td>Завершённых документов (подписаны + архив)</td>
                        <td>{report.docs_completed}</td>
                    </tr>
                    <tr>
                        <td>Подписанных документов</td>
                        <td>{report.docs_signed}</td>
                    </tr>
                    <tr>
                        <td>Завершённых задач workflow</td>
                        <td>{report.tasks_completed}</td>
                    </tr>
                    <tr>
                        <td>Среднее время согласования</td>
                        <td>{avg_days}</td>
                    </tr>
                </tbody>
            </table>

            <div class="footer">
                Сгенерировано: {generated_str} | ГосДок v1.0.0
            </div>
        </body>
        </html>
        """

        pdf_bytes = weasyprint.HTML(string=html_content).write_pdf()
        return pdf_bytes

    except ImportError:
        logger.error("weasyprint не установлен — экспорт PDF недоступен")
        return b""
    except Exception as exc:
        logger.error("Ошибка при генерации PDF: %s", exc)
        return b""
